#!/usr/bin/env python3
"""Build a program KPI waterfall workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


STANDARD = {
    "in_market": 1.0,
    "engage": 0.30,
    "meeting": 0.10,
    "pipeline": 0.50,
    "close": 0.30,
}

PROGRAM_TYPES = [
    "1:1 ABM",
    "1:few ABM",
    "1:many ABM",
    "Executive program",
    "Event follow-up",
    "Digital deal room",
    "Renewal / QBR",
    "Customer expansion",
    "Other",
    "Web Engager Program",
    "Resource Center",
    "Enablement Program",
    "Newsletter",
]
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]

HEADERS = [
    "Program #",
    "+ Program Type",
    "Program Name",
    "Segment / Audience",
    "Channels",
    "Content / Messaging",
    "Notes",
    "Benchmark Mode",
    "Accounts Targeted",
    "Avg Deal Size",
    "Std In-Market %",
    "Std Engage %",
    "Std Meeting %",
    "Std Pipeline %",
    "Std Close %",
    "Custom In-Market %",
    "Custom Engage %",
    "Custom Meeting %",
    "Custom Pipeline %",
    "Custom Close %",
    "Selected In-Market %",
    "Selected Engage %",
    "Selected Meeting %",
    "Selected Pipeline %",
    "Selected Close %",
    "Accounts In Market",
    "Engaged Accounts",
    "Sales Meetings",
    "Pipeline Opps",
    "Closed Deals",
    "Pipeline Goal",
    "Bookings",
    "Quarter",
]


def load_programs(path: str | None) -> list[dict]:
    if not path:
        return [
            {
                "program_type": "1:1 ABM",
                "quarter": "Q1",
                "program_name": "Example Program (replace)",
                "segment": "Target account segment",
                "channels": "Display ads + email + sales outreach",
                "content": "Thought leadership, solution proof, case studies",
                "notes": "Replace this example row with the customer program details.",
                "benchmark_mode": "Standard",
                "accounts_targeted": 1200,
                "avg_deal_size": 5000000,
            }
        ]
    with open(path, "r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError("--programs must be a JSON array")
    return data


def set_col_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in row:
        cell.fill = fill
        cell.font = Font(bold=True, color="1C293F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def build(output: Path, programs: list[dict], rows: int) -> None:
    wb = Workbook()
    form = wb.active
    form.title = "Customer Program Form"
    sections = wb.create_sheet("Waterfall Sections")
    quarter_summary = wb.create_sheet("Quarter Summary")
    bench = wb.create_sheet("Benchmarks")

    blue = PatternFill("solid", fgColor="3C78D8")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    calc_fill = PatternFill("solid", fgColor="F3F9FD")
    border = Border(bottom=Side(style="thin", color="D8ECFA"))

    form.merge_cells("A1:AG1")
    form["A1"] = "Customer Program Waterfall Form"
    form["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    form["A1"].fill = blue
    form["A1"].alignment = Alignment(vertical="center")
    form["A2"] = "Yellow cells are customer inputs. Select a + Program Type and Quarter to build the one-year plan; choose Custom to override the standard benchmark chain."
    form["H2"] = "Standard: 100% in-market, 30% engaged, 10% meeting, 50% pipeline, 30% closed"
    form["A2"].font = Font(italic=True)
    form["H2"].font = Font(italic=True)
    form.append([])
    for col, header in enumerate(HEADERS, 1):
        cell = form.cell(row=3, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.font = Font(bold=True, color="1C293F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    max_rows = max(rows, len(programs) + 25)
    type_validation = DataValidation(type="list", formula1=f'"{",".join(PROGRAM_TYPES)}"', allow_blank=True)
    quarter_validation = DataValidation(type="list", formula1=f'"{",".join(QUARTERS)}"', allow_blank=True)
    mode_validation = DataValidation(type="list", formula1='"Standard,Custom"', allow_blank=False)
    form.add_data_validation(type_validation)
    form.add_data_validation(quarter_validation)
    form.add_data_validation(mode_validation)

    for row in range(4, max_rows + 4):
        idx = row - 4
        p = programs[idx] if idx < len(programs) else {}
        values = [
            f'=IF(B{row}="","",ROW()-3)',
            p.get("program_type"),
            p.get("program_name"),
            p.get("segment"),
            p.get("channels"),
            p.get("content"),
            p.get("notes"),
            p.get("benchmark_mode", "Standard"),
            p.get("accounts_targeted"),
            p.get("avg_deal_size"),
            STANDARD["in_market"],
            STANDARD["engage"],
            STANDARD["meeting"],
            STANDARD["pipeline"],
            STANDARD["close"],
            p.get("custom_in_market"),
            p.get("custom_engage"),
            p.get("custom_meeting"),
            p.get("custom_pipeline"),
            p.get("custom_close"),
            f'=IF($B{row}="","",IF($H{row}="Custom",IF($P{row}="",K{row},$P{row}),K{row}))',
            f'=IF($B{row}="","",IF($H{row}="Custom",IF($Q{row}="",L{row},$Q{row}),L{row}))',
            f'=IF($B{row}="","",IF($H{row}="Custom",IF($R{row}="",M{row},$R{row}),M{row}))',
            f'=IF($B{row}="","",IF($H{row}="Custom",IF($S{row}="",N{row},$S{row}),N{row}))',
            f'=IF($B{row}="","",IF($H{row}="Custom",IF($T{row}="",O{row},$T{row}),O{row}))',
            f'=IF($B{row}="","",ROUND($I{row}*$U{row},0))',
            f'=IF($B{row}="","",ROUND($Z{row}*$V{row},0))',
            f'=IF($B{row}="","",ROUND($AA{row}*$W{row},0))',
            f'=IF($B{row}="","",ROUND($AB{row}*$X{row},0))',
            f'=IF($B{row}="","",ROUND($AC{row}*$Y{row},0))',
            f'=IF($B{row}="","",$J{row}*$AC{row})',
            f'=IF($B{row}="","",$J{row}*$AD{row})',
            p.get("quarter", "Q1") if p else None,
        ]
        for col, value in enumerate(values, 1):
            cell = form.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if 2 <= col <= 10 or 16 <= col <= 20 or col == 33:
                cell.fill = input_fill
            elif col >= 21 or 11 <= col <= 15:
                cell.fill = calc_fill
        type_validation.add(form[f"B{row}"])
        quarter_validation.add(form[f"AG{row}"])
        mode_validation.add(form[f"H{row}"])

    for col in range(11, 26):
        for row in range(4, max_rows + 4):
            form.cell(row=row, column=col).number_format = "0%"
    for row in range(4, max_rows + 4):
        form.cell(row=row, column=9).number_format = "#,##0"
        form.cell(row=row, column=10).number_format = '"$"#,##0'
        for col in range(26, 31):
            form.cell(row=row, column=col).number_format = "#,##0"
        for col in range(31, 33):
            form.cell(row=row, column=col).number_format = '"$"#,##0'

    form.freeze_panes = "A4"
    form.auto_filter.ref = f"A3:AG{max_rows + 3}"
    set_col_widths(form, [12, 18, 24, 28, 30, 32, 36, 17, 16, 16] + [14] * 22 + [12])

    headers = ["Waterfall Metric", "Calculated Value", "Selected Benchmark", "Benchmark / Formula"]
    for block in range(max_rows):
        start = block * 16 + 1
        row_ref = block + 4
        sections.cell(start, 1, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!AG{row_ref}&" Program "&\'Customer Program Form\'!A{row_ref}&": "&\'Customer Program Form\'!B{row_ref}&" - "&\'Customer Program Form\'!C{row_ref})')
        sections.cell(start + 1, 1, "Segment / Audience")
        sections.cell(start + 1, 2, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!D{row_ref})')
        sections.cell(start + 1, 3, "Benchmark Mode")
        sections.cell(start + 1, 4, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!H{row_ref})')
        sections.cell(start + 2, 1, "Channels")
        sections.cell(start + 2, 2, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!E{row_ref})')
        for col, header in enumerate(headers, 1):
            sections.cell(start + 3, col, header)
        rowspec = [
            ("Accounts targeted actively", "I", "", "Customer input"),
            ("Accounts in market", "Z", "U", "Accounts targeted x selected in-market %"),
            ("Engaged accounts", "AA", "V", "Accounts in market x selected engage %"),
            ("Sales meetings scheduled", "AB", "W", "Engaged accounts x selected meeting %"),
            ("Pipeline opportunities", "AC", "X", "Sales meetings x selected pipeline %"),
            ("Closed deals", "AD", "Y", "Pipeline opportunities x selected close %"),
            ("Average deal size", "J", "", "Customer input"),
            ("Pipeline goal", "AE", "", "Average deal size x pipeline opportunities"),
            ("Bookings", "AF", "", "Average deal size x closed deals"),
        ]
        for offset, (metric, value_col, bench_col, note) in enumerate(rowspec, 4):
            sections.cell(start + offset, 1, metric)
            sections.cell(start + offset, 2, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!{value_col}{row_ref})')
            if bench_col:
                sections.cell(start + offset, 3, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!{bench_col}{row_ref})')
            sections.cell(start + offset, 4, note)
        sections.cell(start + 13, 1, "Notes")
        sections.cell(start + 13, 2, f'=IF(\'Customer Program Form\'!B{row_ref}="","",\'Customer Program Form\'!G{row_ref})')
        for row in range(start, start + 14):
            for col in range(1, 5):
                sections.cell(row, col).border = border
                sections.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
        for col in range(1, 5):
            sections.cell(start + 3, col).fill = PatternFill("solid", fgColor="D9D9D9")
            sections.cell(start + 3, col).font = Font(bold=True)
        sections.cell(start, 1).fill = blue
        sections.cell(start, 1).font = Font(bold=True, color="FFFFFF")

    set_col_widths(sections, [32, 22, 20, 42])
    for row in range(1, max_rows * 16 + 1):
        sections.row_dimensions[row].height = 22

    quarter_summary.append(["Quarter", "Programs", "Accounts Targeted", "Sales Meetings", "Pipeline Opps", "Closed Deals", "Pipeline Goal", "Bookings"])
    style_header(quarter_summary[1])
    for idx, quarter in enumerate(QUARTERS, 2):
        quarter_summary.cell(idx, 1, quarter)
        quarter_summary.cell(idx, 2, f'=COUNTIFS(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$B:$B,"<>")')
        quarter_summary.cell(idx, 3, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$I:$I)')
        quarter_summary.cell(idx, 4, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$AB:$AB)')
        quarter_summary.cell(idx, 5, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$AC:$AC)')
        quarter_summary.cell(idx, 6, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$AD:$AD)')
        quarter_summary.cell(idx, 7, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$AE:$AE)')
        quarter_summary.cell(idx, 8, f'=SUMIF(\'Customer Program Form\'!$AG:$AG,A{idx},\'Customer Program Form\'!$AF:$AF)')
    full_row = len(QUARTERS) + 2
    quarter_summary.cell(full_row, 1, "Full Year")
    for col in range(2, 9):
        letter = get_column_letter(col)
        quarter_summary.cell(full_row, col, f"=SUM({letter}2:{letter}{full_row - 1})")
    for row in range(2, full_row + 1):
        for col in range(1, 9):
            quarter_summary.cell(row, col).border = border
            quarter_summary.cell(row, col).alignment = Alignment(wrap_text=True, vertical="center")
        for col in range(7, 9):
            quarter_summary.cell(row, col).number_format = '"$"#,##0'
    set_col_widths(quarter_summary, [14, 12, 18, 18, 18, 16, 18, 18])

    bench.append(["Metric", "Standard Benchmark", "Definition"])
    style_header(bench[1])
    for metric, value, definition in [
        ("Accounts in market", STANDARD["in_market"], "% accounts in market"),
        ("Engaged accounts", STANDARD["engage"], "% conversion from in-market to engaged"),
        ("Sales meetings", STANDARD["meeting"], "% conversion from engaged to sales meeting"),
        ("Pipeline opportunities", STANDARD["pipeline"], "% conversion from sales meeting to pipeline"),
        ("Closed deals", STANDARD["close"], "% conversion from pipeline to deal close"),
    ]:
        bench.append([metric, value, definition])
    for row in range(2, 7):
        bench.cell(row, 2).number_format = "0%"
    set_col_widths(bench, [28, 22, 48])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True, help="Absolute or relative output .xlsx path")
    parser.add_argument("--programs", help="Optional JSON array of program objects")
    parser.add_argument("--rows", type=int, default=100, help="Prepared program rows")
    args = parser.parse_args()
    build(Path(args.output).expanduser().resolve(), load_programs(args.programs), args.rows)


if __name__ == "__main__":
    main()
